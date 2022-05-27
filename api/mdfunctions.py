import pandas as pd
from flask import Response
from math import e
from openpyxl import Workbook
from openpyxl.worksheet.pagebreak import Break
from openpyxl.utils import get_column_letter
from openpyxl.styles import (
    PatternFill,
    Border,
    Side,
    Alignment,
    Protection,
    Font,
    NamedStyle,
)
import xlsxwriter
import csv

from openpyxl.writer.excel import save_virtual_workbook

"""
Function for taking in the values from the sawtooth exported file, and returning the raw utilities and the rescaled utilities

"""
def process_raw_util_xl_file(xl_fn):
    df = pd.read_excel(xl_fn)  # read the excel file
    row = df.shape[0]  # df.shape[0] = Number of rows, df.shape[1] = number of columns
    index_col = [col for col in df if 'id' in col.lower()][0]
    df.set_index(index_col, inplace=True)  # Setting the ID as the index
    # weights = df['Weights'] if 'Weights' in df else pd.Series([1]*row)         # Creating the weights array if there are weights , otherwise an array of 1's with # of rows

    if 'Weights' not in df:
        df['Weights'] = 1


    weights = df['Weights']
    df.drop('Weights', axis=1, inplace=True)
    sum = weights.values.sum()  # Sum of the weights

    # if 'Weights' in df: df = df.drop('Weights', axis=1)                        # removing weights from df if they are there
    rlh_cols = [col for col in df if 'rlh' in col.lower()][:1]
    df = df.drop(rlh_cols, axis=1)  # removing RLH from df if they are there

    max_diff_scores = (e ** df) / (1 + e ** df) * 100  # exp the values of the dataframe with

    utils = df

    return {
        "utilities": utils,
        "claims": [col for col in utils],
        "maxdiff_scores": max_diff_scores,
        "weights": weights
    }

"""
Function to get the reach scores for the items that are ON and Summary Metrics
"""

def calc_reach_metrics(maxdiff_scores, claims_on_list, weights):

    reached = (maxdiff_scores >= 75).astype(int) # converting reached into binary

    reached_items_on = reached[claims_on_list]

    reached_items_on_weighted = reached_items_on.mul(weights, axis=0)  # calculate weighted reach of items that are on
    
    sum_of_weights = weights.values.sum()

    average_reach_percentage_items_on_wtd = (reached_items_on_weighted.sum(axis=0, skipna=True)) / sum_of_weights  # average weighted reach of all the items that are on

    # Count the number of items in the datafile
    number_of_items = reached_items_on.shape[1]

    # items liked average
    num_liked_average_items_on_wtd = (((reached_items_on.sum(axis=1)).mul(weights, axis=0)).sum(axis=0)) / sum_of_weights # average number liked summary metrics weighted

    # calculate the unduplicated reach of the items that are on
    unduplicated_reached_items_on_wtd = (((((reached_items_on.sum(axis=1)) > 0 ).astype(int)).mul(weights, axis=0)).sum(axis=0)) / sum_of_weights
    print(unduplicated_reached_items_on_wtd)

    # calculate the favourite percentage of the items that are on
    favourite_binary = maxdiff_scores.eq(maxdiff_scores.max(axis=1), axis=0).astype(int)

    favourite_binary_on = favourite_binary[claims_on_list]

    fav_scores_items_on = ((favourite_binary_on.mul(weights, axis=0)).sum(axis=0)) / sum_of_weights

    fav_percentage_wtd = fav_scores_items_on.sum(axis=0)

    reach_dict = average_reach_percentage_items_on_wtd.to_dict()
    fav_dict = fav_scores_items_on.to_dict()


    return {
        "Claim_Reach": reach_dict,
        "Claim_Favorite": fav_dict,
        "Summary_Metrics": {
            "Average_Number_of_Items_Liked": num_liked_average_items_on_wtd,
            "Reach": unduplicated_reached_items_on_wtd,
            "Favorite_Percentage": fav_percentage_wtd,
        }
    }

"""
reach function takes in inputs & returns a dictionary with 
    keys: items in order of which they need to be turned on in the simulator, 
    values: metrics in the calc_reach_metrics function
as we go along the keys in the dictionary, values/metrics include the previous items already 'offered'
"""



def get_incremental_reach(max_diff_scores, item_considered, item_on, number_of_items_to_turn_on, weights):
    current_items_to_consider = item_considered.copy()
    current_item_on = item_on.copy()

    get_incremental_reach_summary_metrics = {}  # setting to blank dict

    for i in range(number_of_items_to_turn_on):  # looping for # of items to be added

        steps = {}  # setting steps to be blank dict
        reach_of_items_in_steps = {}  # setting reach_of_items_in_steps to be blank dict
        for test_item in current_items_to_consider:  # looping through all the items that are considered

            new_items_on = [*current_item_on, test_item]  # adding test item to items that are on already

            summary_scores_tested = calc_reach_metrics(max_diff_scores, new_items_on,
                                                       weights)  # calling reach_metric_function


            steps[
                test_item] = summary_scores_tested  # adding the reach_metrics to the steps dict for all the items that are considered, with items name as key

            reach_of_items_in_steps[test_item] = summary_scores_tested['Summary_Metrics'][
                'Reach']  # adding ONLY the unduplicated reach to the reach_of_items_in_steps dict for all the items that are considered, with items name as key

        get_incremental_reach_summary_metrics[max(reach_of_items_in_steps, key=reach_of_items_in_steps.get)] = \
        steps[max(reach_of_items_in_steps,
                  key=reach_of_items_in_steps.get)]  # Item with max reach is key, value is reach metrics of that item
        current_item_on.append(max(reach_of_items_in_steps,
                                   key=reach_of_items_in_steps.get))  # add the item with max reach with item that is on
        current_items_to_consider.remove(max(reach_of_items_in_steps,
                                             key=reach_of_items_in_steps.get))  # remove the items that was found to have the max reach from the considereation

    return {
        "Order of Items": list(get_incremental_reach_summary_metrics.keys()),
        "Incremental Reach Summary": get_incremental_reach_summary_metrics,
    }  # returning the dict for the items that are turned on in order


def process_subgroup_file(xl_fn):
    df_sg = pd.read_excel(xl_fn)

    row = df_sg.shape[0]  # df.shape[0] = Number of rows, df.shape[1] = number of columns
    # index_col = [col for col in df_sg if 'id' in col.lower()][0]
    # df_sg.set_index(index_col, inplace=True)  # Setting the ID as the index
    df_sg.fillna(0)  # handles filling in any missing data in the df
    # weights = df['Weights'] if 'Weights' in df else pd.Series([1]*row)         # Creating the weights array if there are weights , otherwise an array of 1's with # of rows

    if 'responseid' in df_sg: df_sg = df_sg.drop('responseid', axis=1)
    df_sg.set_index('respid', inplace=True)
    if 'Weights' not in df_sg:
        df_sg['Weights'] = 1

    weights = df_sg['Weights']
    df_sg.drop('Weights', axis=1, inplace=True)

    # if 'Weights' in df: df = df.drop('Weights', axis=1)                        # removing weights from df if they are there
    rlh_cols = [col for col in df_sg if 'rlh' in col.lower()][:1]
    df = df_sg.drop(rlh_cols, axis=1)  # removing RLH from df if they are there

    # max_diff_scores = (e ** df) / (1 + e ** df) * 100  # exp the values of the dataframe with

    utils = df

    return {
        "utilities": utils,
        "subgroups": [col for col in utils],
        # "maxdiff_scores": max_diff_scores,
        # "weights": weights
    }


"""
    filter for subgroup using the subgroup config, maxdiff config, and the selected subgroup from the user
"""

def filter_for_subgroup(sg, md, subgroup):
    md = md
    weights = md['weights']
    md_scores = md['maxdiff_scores']
    sg_utils = sg['utilities']
    subgroup = subgroup
    selected_sg = str(subgroup) # convert subgroup to a string
    print(selected_sg)

    # print(weights, md_scores, sg_utils, subgroup )
    filter_by_group = sg_utils.loc[sg_utils[selected_sg] == 1].index.values # filter all the index values (respondent ids) with 1's from the selected sg column
    # print(filter_by_group)

    md_scores_index = md_scores.index.tolist() # convert the mdp config scores to a list
    filtered_respid = [] # all filtered ids that match the condition will be appended to a list
    for id in filter_by_group: # for every respondent id in the subgroup filtered list
        if id in md_scores_index: # if the respondent id in the maxdiff scores & filtered subgroup list match
            filtered_respid.append(id) # then append the id to the filtered respid list
    print(filtered_respid)
    filtered_md_scores = md_scores.loc[filtered_respid] # select the md_scores for only the filtered subgroup respondents
    filtered_weights = weights.loc[filtered_respid] # select the weights for only the filtered subgroup respondents

    return {
        'filtered_md_scores': filtered_md_scores,
        'filtered_weights': filtered_weights,
        'filtered_respid': filtered_respid,
    }


def generate_turf_chart_csv(chart_data):
    items = list(chart_data['orderOfItems'])
    turf_reach = []
    for item in items:
        row = []
        row.append(item)
        turf_reach_in_order = chart_data['incrementalReachSummary'][item]['Summary_Metrics']['Reach']

        percentage_reach = turf_reach_in_order * 100
        # # print(irs[item]['Summary_Metrics']['Reach'])

        round(percentage_reach, 1)
        row.append(str(round(percentage_reach, 1)) + "%")
        turf_reach.append(row)

    print(turf_reach)
    rows = turf_reach

    turf_df = pd.DataFrame(
        rows,
        columns=['Claim', 'Optimized Reach']
    )
    print(turf_df)

    return turf_df

""" function for creating pd df for side by side page csv export """

def generate_prev_sim_csv(data):
    get_dict = data
    claims = data['claims']
    setup_dict = data['setups']
    summary_metric_headers = data['setupSummaryMetricHeaders']
    print(claims)
    print(setup_dict)
    # output = HttpResponse(
    #     mimetype='application/application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = 'Simulation_Summary.xlsx'
    wb = Workbook()
    sheet = wb.active
    # styles
    heading_font = Font(size=11, bold=True)
    heading = NamedStyle(name='Heading')
    wb.add_named_style(heading)
    heading.font = heading_font

    percent_value = NamedStyle(name='Percentage')
    wb.add_named_style(percent_value)
    percent_value.number_format = '0.00%'

    # Claim Header
    headers = ['Claim']
    start_claim_header_row = 1
    start_claim_header_col = 2

    for i, header in enumerate(headers):
        current_row = start_claim_header_row
        column_letter = get_column_letter(start_claim_header_col)
        cell_ref = f"{column_letter}{current_row}"
        sheet[cell_ref] = header
        sheet[cell_ref].style = heading

    # # Setup Header
    # setup_title = "Setup "
    # start_setup_header_row = 1
    # start_setup_header_col = 3
    #
    # for header_index, header in enumerate(setup_dict):
    #     current_row = start_setup_header_row
    #     column_letter = get_column_letter(start_setup_header_col)
    #     cell_ref = f"{column_letter}{current_row}"
    #     sheet[cell_ref] = header_index
    #     sheet[cell_ref].style = heading
    #
    #     for col_index, col_data in enumerate(setup_dict):
    #         current_col = start_setup_header_col + 1
    #         column_letter = get_column_letter(current_col)
    #         cell_ref = f"{column_letter}{current_row}"
    #         sheet[cell_ref] = setup_title + str(col_index)
    #         sheet[cell_ref].style = heading

    # Side by Side Claim and Claim States Table
    starting_col_index = 2
    starting_row_index = 2

    for index, claim in enumerate(claims):
        current_row = starting_row_index + index
        column_letter = get_column_letter(starting_col_index)
        cell_ref = f"{column_letter}{current_row}"
        sheet[cell_ref] = claim
        sheet[cell_ref].style = heading

        for i, setup in enumerate(setup_dict):
            setup_claims_on = setup[3]
            current_col = starting_col_index + i + 1
            column_letter = get_column_letter(current_col)
            cell_ref = f"{column_letter}{current_row}"
            if claim in setup_claims_on:
                sheet[cell_ref] = setup[2][claim]['Summary_Metrics']['Reach']
                sheet[cell_ref].style = percent_value
            elif setup[0][claim] == "Offered":
                sheet[cell_ref] = "Already Offered"
            elif setup[0][claim] == "Considered":
                sheet[cell_ref] = "Considered"
            elif setup[0][claim] == "Excluded":
                sheet[cell_ref] = "Excluded"
            else:
                sheet[cell_ref] = ""

    # Summary Metrics Header
    start_metric_header_row = 16
    start_metric_header_col = 2

    for i, header in enumerate(summary_metric_headers):
        current_row = start_metric_header_row
        column_letter = get_column_letter(start_metric_header_col)
        cell_ref = f"{column_letter}{current_row}"
        sheet[cell_ref] = "Summary Metrics"
        sheet[cell_ref].style = heading

    # Summary Metrics Table

    start_col_index = 2
    start_row_index = 17

    for i, header in enumerate(summary_metric_headers):
        current_row = start_row_index + i
        column_letter = get_column_letter(start_col_index)
        cell_ref = f"{column_letter}{current_row}"
        sheet[cell_ref] = header
        sheet[cell_ref].style = heading

        for id, setup in enumerate(setup_dict):
            current_col = starting_col_index + id + 1
            column_letter = get_column_letter(current_col)
            cell_ref = f"{column_letter}{current_row}"
            if header == "Subgroup":
                sheet[cell_ref] = setup[5]
            elif header == "Number of Respondents":
                sheet[cell_ref] = setup[4]
            elif header == "Average Liked":
                sheet[cell_ref] = round(setup[1]["Average_Number_of_Items_Liked"], 2)
            elif header == "Average Reach":
                sheet[cell_ref] = setup[1]["Reach"]
                sheet[cell_ref].style = percent_value
            elif header == "Average Favorite":
                sheet[cell_ref] = setup[1]["Favorite_Percentage"]
                sheet[cell_ref].style = percent_value
            else:
                sheet[cell_ref] = ""

    wb.save(filename=filename)
    print(wb)
    return save_virtual_workbook(wb)

# Response(
#         save_virtual_workbook(wb),
#         headers={
#             'Content-Disposition': 'attachment; filename=sheet.xlsx',
#             'Content-type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
#         }
#     )








